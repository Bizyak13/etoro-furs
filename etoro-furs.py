# !/usr/bin/env python3
# -*- coding: utf-8 -*-

import re
import csv
import json
import os.path
import argparse
import requests
import xml.etree.ElementTree
from datetime import datetime
from datetime import timedelta
import configparser

import openpyxl


ENCODING = 'utf-8'
DF_ETORO = '%d/%m/%Y'
DF_FURS = '%d. %m. %Y'
DF_XML = '%Y-%m-%d'

CURRENCY_FILE = 'currency-rates.xml'

headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
}

bsi_rate_url = 'https://www.bsi.si/_data/tecajnice/dtecbs-l.xml'

csv_1st_line = '#FormCode;Version;TaxPayerID;TaxPayerType;DocumentWorkflowID;;;;;;'
csv_2nd_line = 'DOH-DIV;3.9;xxxxxxxx;FO;O;;;;;;'
csv_3rd_line = '#datum prejema dividende;davčna številka izplačevalca dividend;identifikacijska  številka izplačevalca dividend;naziv izplačevalca dividend;naslov izplačevalca dividend;država izplačevalca dividend;vrsta dividende;znesek dividend;tuji davek;država vira;uveljavljam oprostitev po mednarodni pogodbi'


def get_rounded_float(number) -> str:
    """
    Get a string representation of a number rounded to two decimal places. Takes into account the format and if there is a thousands separator present.

    Parameters
    ----------
    number: Any
        Number to transform
        
    Returns
    -------
    str
        Representation of a given number in .2f format with ',' delimiter
    """
    number = str(number)
    if ',' in number and '.' in number:
        if number.index(',') < number.index('.'):
            number = number.replace(',', '')
        else:
            number = number.replace('.', '')            
    number = float(number)
    number = f'{number:.2f}'
    return str(number).replace('.', '#').replace(',', '.').replace('#', ',')


def get_file_validity(file, days) -> bool:
    """
    Get the validity of file age (modified). Basically checks if the file is older than 'days'

    Parameters
    ----------
    file: str
        Location of the file in string format
    days: int
        Number of days to go back to
        
    Returns
    -------
    bool
        Boolean if file age passed the check or not
    """
    modified_date = datetime.fromtimestamp(os.path.getmtime(file))
    days_ago = datetime.today() - timedelta(days=days)
    if modified_date < days_ago:
        return True
    return False


def remove_offending_rows(table, keep, col) -> openpyxl.worksheet:
    """
    Removes offending rows from the xlsx file that are not needed

    Parameters
    ----------
    table: openpyxl.worksheet
        The worksheet with offending rows
    keep: str
        The string of cell value to check if row is affected
    col: int
        The column location for the check
        
    Returns
    -------
    openpyxl.worksheet
        Returns the cleaned worksheet object with rows removed
    """
    for row in range(table.max_row, 1, -1):
        if table.cell(row=row, column=col).value != keep:
            table.delete_rows(row, 1)
    
    return table


def get_config() -> configparser.RawConfigParser:
    """
    Get the config object from parsing the config file
    """
    config = configparser.RawConfigParser()
    config.read('CONFIG.cfg')
    return config


def get_config_value(section, key) -> any:
    """
    Get config value from the config

    Parameters
    ----------
    section: str
        Get the section in the config
    key: str
        Get the key in the section of the config
        
    Returns
    -------
    any
        Returns the config value of the section.key
    """
    value = config[section][key]
    if value is not None:
        return config[section][key]
    else: 
        raise SystemExit(f'ERROR: Missing config value for {section}:{key}')
        


def parse_args() -> list:
    parser = argparse.ArgumentParser()
    parser.add_argument('input', help='Input file from etoro. Must be in .xlsx format.', type=str)
    parser.add_argument('output', help='Output file from csv. Must be in .csv format.', type=str)
    parser.add_argument('-v', '--verbose', help='Verbose output', action='store_true')
    args = parser.parse_args()

    if args.input.split('.')[-1] != 'xlsx':
        raise SystemExit('ERROR: Provided input file not in the right format')

    if os.path.exists(args.input) == False:
        raise SystemExit('ERROR: Provided input file does not exist')

    return args


def get_conversion_rate_file() -> dict:
    if not os.path.exists(CURRENCY_FILE) or get_file_validity(CURRENCY_FILE, 2):
        response = requests.get(bsi_rate_url, headers=headers)
        if response.status_code == 200:
            print("etoro-furs: XML file downloaded successfully.")
            with open(CURRENCY_FILE, "wb") as file:
                file.write(response.content)
        else:
            print("etoro-furs: Currency rates file could not be downloaded.")
            exit(1)

    conversion_file = xml.etree.ElementTree.parse(CURRENCY_FILE).getroot()

    rates = {}
    for d in conversion_file:
        date = d.attrib["datum"].replace("-", "")
        rates[date] = {}
        for r in d:
            currency = r.attrib["oznaka"]
            rates[date][currency] = r.text
    return rates


def get_conversion_rate_on_date(rates, date, currency) -> any:
    date = date.strftime('%Y%m%d')
    if date in rates and currency in rates[date]:
        rate = rates[date][currency]
    else:
        rate = 0
    return rate


def get_config_taxid() -> int:
    tax_id = get_config_value('TAX_ID','tax_id')
    if len(tax_id) == 8 and int(tax_id):
        return tax_id
    else:
        raise SystemExit('ERROR: Tax ID is not in the right format')


def parse_input_file(rates) -> dict:
    data = {}
    workbook = openpyxl.load_workbook(args.input)
    dividends = workbook['Dividends']
    activity = workbook['Account Activity']
    activity = remove_offending_rows(activity, 'Dividend', 2)

    max_col = dividends.max_column
    max_row = dividends.max_row

    for i in range(2, max_row+1):
        rate = 0
        data_row = {}
        date = datetime.strptime(dividends.cell(row=i, column=1).value, DF_ETORO)
        # get price data for every row, based on the date
        for j in range(1, max_col + 1):
            data_row[dividends.cell(row=1, column=j).value] = dividends.cell(row=i, column=j).value
            data_row['Date of Payment FURS'] = date.strftime(DF_FURS)
            if j == 2:
                # get data about the company and append it here
                symbol_curr = activity.cell(row=i, column=3).value.split('/')
                data_row['Symbol'] = symbol_curr[0]
                data_row['Currency'] = symbol_curr[1]
                company_json = json.load(open('companies.json', encoding=ENCODING))
                for company in company_json['companies']:
                    if company['symbol'] == activity.cell(row=i, column=3).value.split('/')[0]:
                        data_row['Company Name'] = company['name']
                        data_row['Company Address'] = company['address']
                        data_row['Company Country'] = company['country']
                        data_row['Company TAX ID'] = company['taxNumber']
                
                if not 'Company Name' in data_row:
                    print(f'etoro-furs: Company <{activity.cell(row=i, column=3).value.split("/")[0]}> not found in companies.json, please add it!')
                    exit(1)

            net_col = 3
            tax_col = 10
            if j == net_col:
                if date.strftime('%Y%m%d') in rates:
                    rate = float(rates[date.strftime('%Y%m%d')][activity.cell(row=i, column=3).value.split('/')[1]])
                else:
                    for k in range(1, 10):
                        last_working = (date - timedelta(days=k)).strftime('%Y%m%d')
                        if last_working in rates:
                            rate = float(rates[last_working][data_row['Currency']])
                            data_row['Conversion rate date'] = (date - timedelta(days=k)).strftime(DF_XML)
                            break
                    if rate == 0:
                        raise SystemExit('ERROR: No exchange rate found for this date')
                data_row[f'Conversion rate (EUR/{data_row["Currency"]})'] = rate
                gross_script = (dividends.cell(row=i, column=net_col).value/rate + dividends.cell(row=i, column=tax_col).value/rate)
                gross_etoro = (dividends.cell(row=i, column=net_col+1).value + dividends.cell(row=i, column=tax_col+1).value)
                
                if args.verbose:
                    print(f'Scrpt € | net: {dividends.cell(row=i, column=net_col).value/rate:.4f},\ttax: {dividends.cell(row=i, column=tax_col).value/rate:.4f},\tgross: {gross_script:.4f}\trate: {rate:.4f}')
                    print(f'Etoro € | net: {dividends.cell(row=i, column=net_col+1).value:.4f},\ttax: {dividends.cell(row=i, column=tax_col+1).value:.4f},\tgross: {gross_etoro:.4f}\trate: {(dividends.cell(row=i, column=tax_col).value/dividends.cell(row=i, column=tax_col+1).value):.4f}')
                    print(f'================================================================')
                
                data_row['Gross Dividend Received (EUR) script'] = get_rounded_float(gross_script)
                data_row['Gross Dividend Received (EUR) etoro'] = get_rounded_float(gross_etoro)
            
            if j == tax_col:
                data_row['Withholding Tax Amount (EUR) script'] = get_rounded_float(dividends.cell(row=i, column=tax_col).value/rate)
                data_row['Withholding Tax Amount (EUR) etoro'] = get_rounded_float(dividends.cell(row=i, column=tax_col+1).value)
        
        data[i-1] = data_row
    return data


def create_output_file(data) -> str:
    
    if 'csv' not in args.output:
        output = args.output + '.csv'
    else:
        output = args.output
    
    csv_output = open(output, 'w', encoding=ENCODING)
    csv_writer = csv.writer(csv_output, lineterminator='\n', delimiter=';')
    csv_writer.writerow(csv_1st_line.split(';'))
    csv_writer.writerow(csv_2nd_line.replace('xxxxxxxx', get_config_taxid()).split(';'))
    csv_writer.writerow(csv_3rd_line.split(';'))
    for l in range(1, len(data)+1):
        line = data[l]
        dividend_type = get_config_value('TAX_ID', 'DIVIDEND_TYPE')
        if not dividend_type:
            dividend_type = 1
        csv_writer.writerow([line['Date of Payment FURS'], '', line['Company TAX ID'], line['Company Name'], line['Company Address'], line['Company Country'], dividend_type, line['Gross Dividend Received (EUR) script'], line['Withholding Tax Amount (EUR) script'], line['Company Country'], ''])
    csv_output.close()

    return output


if __name__ == '__main__':
    print('etoro-furs: Running etoro-furs')
    args = parse_args()
    config = get_config()
    rates = get_conversion_rate_file()
    print('etoro-furs: Rates loaded')
    
    print('etoro-furs: Parsing input file')
    data = parse_input_file(rates)
    if args.verbose:
        print(json.dumps(data))
    print('etoro-furs: Generating csv file')
    file = create_output_file(data)
    print(f'etoro-furs: DONE! -> {file}')
